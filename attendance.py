#anil gaikwad
from tkinter import *
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

root = Tk()
root.title('Attendence Student')
root.geometry("500x600")

wb = Workbook()

wb = load_workbook('attendance.xlsx')
ws=wb.active

column_a=ws['B']
column_b=ws['C']

def get_a():
    List=''
    for cell in column_a:
        List = f'{List + str(cell.value)}\n\n'
    label_a.config(text=List)
    List1 = ''
    for cell in column_b:
        List1 = f'{List1 + str(cell.value)}\n\n'
    label_b.config(text=List1)



b1=Button(root,text="VIEW" ,command = get_a)
b1.pack(pady=20)
label_a = Label (root ,text="")
label_a.pack(pady=20)
label_b = Label (root ,text="")
label_b.pack(pady=20)

root.mainloop()