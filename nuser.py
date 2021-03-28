#anil gaikwad
import os
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
import datetime
import time
import sqlite3
from openpyxl import load_workbook
import dlib
from skimage import io

from subprocess import call
speech="Welcome new student!"
call(["espeak",speech])

def quit(*args):
    root.destroy()
##database
def database():
    roll = t.get()
    name = t1.get()
    email = t2.get()
    wb = load_workbook(f'data/excel/data.xlsx')

    sheet = wb.active
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 20

    sheet.cell(row=1, column=1).value = "Roll Number"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Email Id"

    current_row = sheet.max_row
    current_column = sheet.max_column

    sheet.cell(row=current_row + 1, column=1).value = t.get()
    sheet.cell(row=current_row + 1, column=2).value = t1.get()
    sheet.cell(row=current_row + 1, column=3).value = t2.get()

    wb.save(f'data/excel/data.xlsx')

    global conn, cursor
    conn = sqlite3.connect('students.db')
    cursor = conn.cursor()
    try:
        cursor.execute('''CREATE TABLE  students1(roll_numbers integer PRIMARY KEY, names text not null, emails text );''')

        if cursor:
            cursor.execute("INSERT INTO students1(roll_numbers,names,emails) values ('" + roll + "','"+ name + "','"+  email +"')")
            print("inserted")
            speech = "successfully Registered"
            call(["espeak", speech])
    except sqlite3.Error as error:
        print("Failed to insert data into sqlite table", error)
    finally:
            conn.commit()
            cursor.close()
            print("Sqlite3 connection closed")


    t.delete(0, 'end')
    t1.delete(0, 'end')
    t2.delete(0, 'end')

#######
root = Tk()
root.attributes("-fullscreen", True)
######

class Example(Frame):
    def __init__(self, master, *pargs):
        Frame.__init__(self, master, *pargs)

        self.image = Image.open("backimage/back.png")
        self.img_copy= self.image.copy()

        self.background_image = ImageTk.PhotoImage(self.image)

        self.background = Label(self, image=self.background_image)
        self.background.pack(fill=BOTH, expand=YES)
        self.background.bind('<Configure>', self._resize_image)

    def _resize_image(self,event):

        new_width = event.width
        new_height = event.height

        self.image = self.img_copy.resize((new_width, new_height))

        self.background_image = ImageTk.PhotoImage(self.image)
        self.background.configure(image =  self.background_image)
e = Example(root)
e.pack(fill=BOTH, expand=YES)
#C = tk.Canvas(root, height=600, width=600)


ts = time.time()
date = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
day,month,year=date.split("-")

mont={'01':'January',
      '02':'February',
      '03':'March',
      '04':'April',
      '05':'May',
      '06':'June',
      '07':'July',
      '08':'August',
      '09':'September',
      '10':'October',
      '11':'November',
      '12':'December'
      }

message3 = tk.Label(root, text="Smart Attendance System" ,fg="white",bg="#8a2e7f" ,width=55 ,height=1, font=('times new roman', 30, ' bold '))
message3.place(relx = 0.5, rely = 0.1, anchor = CENTER)

frame1 = tk.Frame(root, bg="#8a2e7f")
frame1.place(relx=0.5, rely=0.17, relwidth=0.2, relheight=0.05,anchor=CENTER)

datef = tk.Label(frame1, text = day+"-"+mont[month]+"-"+year, fg="white",bg="#2B3856" ,width=55 ,height=1,font=('times new roman', 20, ' bold '))
datef.pack(fill='both',expand=1)
##
frame2 = tk.Frame(root, bg="#ffffff")
frame2.place(relx=0.50, rely=0.56, relwidth=0.35, relheight=0.70 ,anchor=CENTER)


head2 = tk.Label(frame2, text="New Student    ", width=40 , fg="white",bg="brown" ,font=('times new roman', 18, ' bold ') )
head2.grid(row=0,column=0)

###data
lbl = tk.Label(frame2, text="Enter Roll Number",width=20  ,height=1  ,fg="white"  ,bg="#8a2e7f" ,font=('times new roman', 16, ' bold '))
lbl.place(relx=0.3, rely=0.1)

txt = tk.Entry(frame2,width=16 ,fg="black",font=('times new roman', 15 ))
t=Entry(frame2,width="28")
t.place(relx=0.32, rely=0.16)
#
lbl2 = tk.Label(frame2, text="Enter Student Name",width=20,height=1  ,fg="white"  ,bg="#8a2e7f" ,font=('times new roman', 16, ' bold '))
lbl2.place(relx=0.3, rely=0.27)
#
txt2 = tk.Entry(frame2,width=20 ,fg="black",font=('times new roman', 15))
t1=Entry(frame2,width="28")
t1.place(relx=0.3, rely=0.33)
#
lbl3 = tk.Label(frame2, text="Enter Email Id",width=20,height=1   ,fg="white"  ,bg="#8a2e7f" ,font=('times new roman', 16, ' bold '))
lbl3.place(relx=0.3, rely=0.42)

txt3= tk.Entry(frame2,width=20 ,fg="black",font=('times new roman', 15)  )
t2=Entry(frame2,width="28")
t2.place(relx=0.3, rely=0.49)
#camera
def cam():
    os.system('python3 take_image.py')
Img = tk.Button(frame2, text="Take Images",fg="white" ,command= cam  ,bg="#507d2a" , width=14  ,height=1, activebackground = "white" ,font=('times new roman', 15, ' bold '))
Img.place(relx=0.1,rely=0.60)
# model save
def train():
   os.system('python3 extraction.py')
Img1 = tk.Button(frame2, text="Train Model",fg="white" ,command= train  ,bg="#507d2a" , width=14  ,height=1, activebackground = "white" ,font=('times new roman', 15, ' bold '))
Img1.place(relx=0.5,rely=0.60)
###submit
Img2 = tk.Button(frame2, text="Submit",fg="white"  ,bg="#254117" , width=14  ,height=1, command= database , activebackground = "white" ,font=('times new roman', 15, ' bold '))
Img2.place(relx=0.3,rely=0.80)
##exit
r = Button(root, text="EXIT", bd=10, command=quit, font=('times new roman', 12), bg="green", fg="yellow", height=1, width=8).place(x=1037, y=700)
##main
if __name__ == '__main__':

    root.mainloop()

##end
