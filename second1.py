##@nil gaikwad
import tkinter
import os
import tkinter as tk
from datetime import time, datetime
from tkinter import *
from PIL import ImageTk, Image
import datetime
import time
import sqlite3
import cv2
from openpyxl import load_workbook

from subprocess import call
speech="Authentication Successful ,Welcome!"
call(["espeak",speech])

def quit(*args):
    os.system('python3 mail.py')
    root.destroy()
##database
def database():
    roll = t.get()
    name = t1.get()
    Att = 'Present'
    ts = time.time()
    date = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
    timeStamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')

    wb = load_workbook(f'attendance.xlsx')

    sheet = wb.active
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

    sheet.cell(row=1, column=1).value = "Roll Number"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Attendance"
    sheet.cell(row=1, column=4).value = "Date"
    sheet.cell(row=1, column=5).value = "Time"

    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row + 1, column=1).value = t.get()
    sheet.cell(row=current_row + 1, column=2).value = t1.get()
    sheet.cell(row=current_row + 1, column=3).value = Att
    sheet.cell(row=current_row + 1, column=4).value = date
    sheet.cell(row=current_row + 1, column=5).value = timeStamp

    wb.save(f'attendance.xlsx')

    global conn, cursor
    conn = sqlite3.connect('students.db')
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS student(roll_number integer  PRIMARY KEY, name text not null);''')
    cursor.execute("SELECT * FROM students1 WHERE 'roll_numbers' ='" + roll + "' AND 'names' ='" + name +"'")
    if cursor:
        print("data inserting")
        cursor.execute("INSERT INTO student (roll_number,name) values ('" + roll + "','"+ name + "');")
        print("inserted successfully")
        speech = "Authentication match"
        call(["espeak", speech])
       # conn.commit()
        #cursor.close()

def Login():
    database()
    if t.get() == "" or t1.get() == "":
       print("Do not match")
       speech = "Please complete the required field "
       call(["espeak", speech])
    else:
        cursor.execute("SELECT * FROM students1 WHERE `roll_numbers` = ? AND `names` = ?",
                           (t.get(), t1.get()))
        if cursor.fetchone() is not None:
            t.delete(0, 'end')
            t1.delete(0, 'end')
            print("Inserted")
            speech = "Attendance successfully Submited "
            call(["espeak", speech])
        else:
            print("Do not match")
            speech = "Invalid rollnumber and name "
            call(["espeak", speech])

            t.delete(0, 'end')
            t1.delete(0, 'end')
    #database()
    conn.commit()
    cursor.close()

##
root = Tk()
root.attributes("-fullscreen", True)
root.configure(background='black')
root.bind("<Escape>", quit)
root.bind("x", quit)
#####using class

class Example(Frame):
    def __init__(self, master, *pargs):
        Frame.__init__(self, master, *pargs)

        self.image = Image.open("backimage/back1.jpg")
        self.img_copy= self.image.copy()

        self.background_image = ImageTk.PhotoImage(self.image)

        self.background = Label(self, image=self.background_image)
        self.background.pack(fill=BOTH, expand=YES)
        self.background.bind('<Configure>', self._resize_image)

    def _resize_image(self,event):
        new_width = event.width
        new_height = event.height
        self.image = self.img_copy.resize((new_width, new_height))
        self.background_image = ImageTk.PhotoImage(self.image)
        self.background.configure(image =  self.background_image)
e = Example(root)
e.pack(fill=BOTH, expand=YES)
#C = tk.Canvas(root, height=600, width=600)

ts = time.time()
date = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y')
day,month,year=date.split("-")

mont={'01':'January',
      '02':'February',
      '03':'March',
      '04':'April',
      '05':'May',
      '06':'June',
      '07':'July',
      '08':'August',
      '09':'September',
      '10':'October',
      '11':'November',
      '12':'December'
      }
#####Main page

message3 = tk.Label(root, text="Smart Attendance System" ,fg="white",bg="#8a2e7f" ,width=55 ,height=1,font=('arial', 30, ' bold '))
message3.place(relx = 0.5, rely = 0.1, anchor = CENTER)

frame4 = tk.Frame(root, bg="#8a2e7f")
frame4.place(relx=0.5, rely=0.17, relwidth=0.2, relheight=0.05,anchor=CENTER)

datef = tk.Label(frame4, text = day+"-"+mont[month]+"-"+year, fg="white",bg="#2B3856" ,width=55 ,height=1,font=('arial', 20, ' bold '))
datef.pack(fill='both',expand=1)

frame1 = tk.Frame(root, bg="#ffffff")
frame1.place(relx=0.35, rely=0.21, relwidth=0.35, relheight=0.70)

head1 = tk.Label(frame1, text="Students Register  ",width=40, fg="white",bg="brown" ,font=('arial', 18, ' bold ') )
head1.place(x=0,y=0)

lbl = tk.Label(frame1, text="Enter Roll Number",width=20  ,height=1  ,fg="white"  ,bg="#8a2e7f" ,font=('arial', 16, ' bold '))
lbl.place(relx=0.2, rely=0.1)

txt = tk.Entry(frame1,width=16 ,fg="black",font=('arial', 15 ))
t=Entry(frame1,width="28")
t.place(relx=0.2, rely=0.16)
#
lbl2 = tk.Label(frame1, text="Enter Student Name",width=20,height=1  ,fg="white"  ,bg="#8a2e7f" ,font=('arial', 16, ' bold '))
lbl2.place(relx=0.2, rely=0.27)
#
txt2 = tk.Entry(frame1,width=20 ,fg="black",font=('arial', 15))
t1=Entry(frame1,width="28")
t1.place(relx=0.2, rely=0.33)

def recg():
    os.system('python3 facerecog.py')
Img2 = tk.Button(frame1, text="Recognize Image",fg="white" ,command = recg , bg="#3090C7" , width=24  ,height=1, activebackground = "white" ,font=('arial', 15, ' bold '))
Img2.place(relx=0.2,rely=0.42)

###
def addstd():
    os.system('python3 nuser.py')
Img1 = tk.Button(frame1, text="New Student",fg="white",command=addstd  , bg="#507d2a" , width=14  ,height=1, activebackground = "white" ,font=('arial', 15, ' bold '))
Img1.place(relx=0.56,rely=0.80)
###
def at():
   os.system('python3 attendance.py')
b3 = tk.Button(frame1, text="View Attendance",fg="white" , command = at , bg="#507d2a" , width=14  ,height=1, activebackground = "white" ,font=('arial', 15, ' bold '))
b3.place(relx=0.1,rely=0.80)
###
b4 = tk.Button(frame1, text="Submit Attendance",fg="white",command =Login , bg="#3090C7" , width=24  ,height=1, activebackground = "white" ,font=('arial', 15, ' bold '))
b4.place(relx=0.2,rely=0.54)
##
r = Button(root, text="EXIT", bd=10,
           command=quit, font=('arial', 12), bg="green", fg="yellow", height=1, width=8).place(x=1037, y=700)
###
if __name__ == '__main__':
    root.mainloop()


##end
